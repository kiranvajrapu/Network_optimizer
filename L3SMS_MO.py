import pandas as pd
import Dataset
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Color,fills,Side,Border,GradientFill,Font,Alignment

def L3SMS_MO():
    # importing input file
    ds = Dataset.dataset_extract()
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    #drop unnamed columns and extra columns
    ds = ds.drop(['EQ', 'Frame Number', 'EventInfo'], axis=1)
    ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

    #rename columns names at pandas
    ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "Serving Cell DL EARFCN",
                                       "All-Serving Cell Identity[1]": "Serving Cell Identity"})
    ds = ds[(ds["Serving Cell DL EARFCN"] != 1400)]
    ds = ds[(ds['Direction'] == 'DL')]
    ds = ds[(ds['Message Type'] == 'CP-Ack') | (ds['Message Type'] == 'CP-Data') ]
    ### removes columns time data
    ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

    ### attach delay column

    #### droping duplicates
    s2 = ds['Message Type']
    # print(s1.values.tolist())
    t2 = s2.values.tolist()
    # print(t1[-1])
    l2 = []
    for i in range(0, len(t2) - 1):
        if t2[i] == t2[i + 1]:
            # print('equal')
            if t2[i] == 'CP-Ack':
                l2.append(i)
            else:
                l2.append(i + 1)
        # else:
        # print("not equal")
    if t2[-1] == 'CP-Ack':
        l2.append(len(t2) - 1)
    # print(l1)
    ds = ds.drop(ds.index[l2])
    ##########TO DROP REMAINING SMS SENTS##########
    ev = ds['Event']
    #print(ev)
    ev1 = ev.values.tolist()
    evl = []
    for i in range(0, len(ev1)):
        if ev1[i] == 'SMS Sent':
            #print(i)
            evl.append(i - 1)
            evl.append(i)
    #print(evl)
    ds = ds.iloc[evl]

    wb = openpyxl.load_workbook('sample layer 3 format.xlsx')
    ws = wb.get_sheet_by_name('L3SMS MO')
    #ws = wb.active

    redFill = PatternFill(start_color=Color('FFFF00'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    #print(ws)
    for r in dataframe_to_rows(ds, index=False, header=False):
        ws.append(r)

    # ws.delete_rows(1, amount=1)

    # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
    #    for cell in rows:
    #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)






    row_count = ws.max_row
    #print(row_count)


    for i in range(3, row_count + 1, 2):
        index = "D" + str(i)
        ws[index].fill = yellowFill

    for i in range(2, row_count + 1):
        ws["A" + str(i)].border = thin_border
        ws["B" + str(i)].border = thin_border
        ws["C" + str(i)].border = thin_border
        ws["D" + str(i)].border = thin_border
        ws["E" + str(i)].border = thin_border
        ws["F" + str(i)].border = thin_border

    wb.save('sample layer 3 format.xlsx')

