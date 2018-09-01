import pandas as pd
import openpyxl
import Dataset
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Color,fills,Side,Border,GradientFill,Font,Alignment

def L3_ERAB_ESTABLISH():
    # importing input file

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ds = Dataset.dataset_extract()

    #drop unnamed columns and extra columns
    ds = ds.drop(['EQ', 'Frame Number', 'Event', 'EventInfo'], axis=1)
    ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

    #rename columns names at pandas
    ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                       "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
    ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]
    ds = ds[(ds['Message Type'] == 'Activate Default EPS Bearer Context Request') | (ds['Message Type'] == 'Activate Default EPS Bearer Context Accept')]
    ### removes columns time data
    ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)



    #### duplicate remover
    s = ds['Message Type']
    #print(s.values.tolist())
    t = s.values.tolist()
    #print(t[-1])
    l = []
    for i in range(0, len(t) - 1):
        if t[i] == t[i + 1]:
            #print('equal')
            l.append(i)
    if t[-1] == 'Activate Default EPS Bearer Context Request':
        l.append(len(t) - 1)
    # print(l)
    ds = ds.drop(ds.index[l])



    #### openpyxl mode



    wb = openpyxl.load_workbook('sample layer 3 format.xlsx')
    ws = wb.get_sheet_by_name('L3-ERAB ESTABLISH')
    #ws = wb.active
    size_of_pandas = len(ds['Time']) + 1
    redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    #print(ws)
    for r in dataframe_to_rows(ds, index=False, header=False):
        ws.append(r)

    # ws.delete_rows(1, amount=1)

    # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
    #    for cell in rows:
    #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

    ### side_table


    ws['H3'] = "ERAB Establish Success Rate"
    ws.merge_cells('H3:I3')
    ws['H3'].alignment = Alignment(horizontal='center')
    ws['H3'].fill = yellowFill
    ### Fields
    ws['H4'] = "Activate Default EPS Bearer Context Request"
    ws['H5'] = "Activate Default EPS Bearer Context Accept"
    ws['H6'] = "EPS Bearer Success rate"


    ws['H3'].border = thin_border
    ws['H4'].border = thin_border
    ws['H5'].border = thin_border
    ws['H6'].border = thin_border

    row_count = ws.max_row
    #print(row_count)


    for i in range(3, row_count + 1, 2):
        index = "C" + str(i)
        ws[index].fill = redFill


    wb.save('sample layer 3 format.xlsx')

    #ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='L3-ERAB ESTABLISH')



    count_Request = 0
    for i in ds['Message Type']:
        if i == "Activate Default EPS Bearer Context Request":
            count_Request = count_Request + 1

    count_Complete = 0
    for i in ds['Message Type']:
        if i == "Activate Default EPS Bearer Context Accept":
            count_Complete = count_Complete + 1

    if count_Request==count_Complete:
        success_rate = str(100) + "%"
    else:
        success_rate = str(99) + "%"

    ws['I4'] = count_Request
    ws['I5'] = count_Complete
    ws['I6'] = success_rate


    ws['I4'].border = thin_border
    ws['I5'].border = thin_border
    ws['I6'].border = thin_border

    for i in range(2, row_count + 1):
        ws["A" + str(i)].border = thin_border
        ws["B" + str(i)].border = thin_border
        ws["C" + str(i)].border = thin_border
        ws["D" + str(i)].border = thin_border
        ws["E" + str(i)].border = thin_border


    wb.save('sample layer 3 format.xlsx')


