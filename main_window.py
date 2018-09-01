import sys
from sys import exit
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton,QProgressBar ,QMainWindow,QInputDialog, QLineEdit, QFileDialog,QLabel,QStatusBar,QComboBox
import openpyxl
import Dataset

from PyQt5.QtGui import *
import pandas as pd
import numpy as np
import openpyxl
import Dataset
from shutil import copyfile


from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Color,fills,Side,Border,GradientFill,Font,Alignment
# adding exception handling


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
class App(QWidget):



    def __init__(self):
        super().__init__()
        self.title = 'Cyient'
        self.progressBar = QProgressBar(self)
        self.progressBar.setGeometry(100,400,200,25)
        #self.statusBar = QStatusBar
        self.left = 10
        self.top = 10
        self.width = 500
        self.height = 500
        self.initUI()




    def initUI(self):
        self.setWindowTitle(self.title)
        self.setWindowIcon(QIcon("icon.png"))
        self.setGeometry(self.left, self.top, self.width, self.height)
        #self.statusBar().showMessage('Message in statusbar.')
        self.sourcefile = QLabel(self)
        self.ul_file = QLabel(self)
        self.dl_file = QLabel(self)
        self.output = QLabel(self)
        self.output_lable = QLabel(self)
        self.textbox = QLineEdit(self)
        self.textbox.move(170, 20)
        self.textbox.resize(100, 20)
        self.sourcefile.move(50,150)
        self.ul_file.move(50,170)
        self.dl_file.move(50,190)
        self.output.move(50,210)
        self.output_lable.move(50,20)
        self.output_lable.setText("Enter Output file name")



        FMT_button = QPushButton('Select input file', self)
        FMT_button.setToolTip('This is an example button')
        FMT_button.move(50, 50)
        FMT_button.clicked.connect(self.openFileNameDialog)
        UL_button = QPushButton('Select UL file', self)
        UL_button.setToolTip('Select DL data file')
        UL_button.move(150, 50)
        UL_button.clicked.connect(self.openFileNameDialog_UL)
        DL_button = QPushButton('Select DL file', self)
        DL_button.setToolTip('Select DL data file ')
        DL_button.move(250, 50)
        DL_button.clicked.connect(self.openFileNameDialog_DL)
        Compute = QPushButton('Automate', self)
        Compute.setToolTip('Click to automate')
        Compute.move(50, 100)
        Compute.clicked.connect(self.process)
        set = QPushButton('set name', self)
        set.setToolTip('set name')
        set.move(300, 20)
        set.clicked.connect(self.set_output_file_name)


        self.show()


    def openFileNameDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        global filename
        filename, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "",
                                                  "All Files (*);;Python Files (*.py)", options=options)
        self.sourcefile.setText("Source File :"+filename)
        self.sourcefile.adjustSize()


    def openFileNameDialog_UL(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        global filename_UL
        filename_UL, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "",
                                                  "All Files (*);;Python Files (*.py)", options=options)
        self.ul_file.setText("UL File :"+filename_UL)
        self.ul_file.adjustSize()



    def openFileNameDialog_DL(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        global filename_DL
        filename_DL, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "",
                                                      "All Files (*);;Python Files (*.py)", options=options)
        self.dl_file.setText("DL File :"+filename_DL)
        self.dl_file.adjustSize()




    def set_output_file_name(self):
        filename_temp = self.textbox.text()
        global source
        global target
        global filename_output
        source = "empty sheet/sample layer 3 format.xlsx"
        target = "Output/" + filename_temp + ".xlsx"
        try:
            copyfile(source, target)
        except IOError as e:
            print("Unable to copy file. %s" % e)
            exit(1)
        except:
            print("Unexpected error:", sys.exc_info())
            exit(1)

        print("\nFile copy done!\n")
        filename_output = target



    global l3_ATTACH_SR_WITH_DELAY_CALCULA
    def l3_ATTACH_SR_WITH_DELAY_CALCULA(self):

        ds = pd.read_excel(filename)

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'Event', 'EventInfo'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
        ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]
        ds = ds[(ds['Message Type'] == 'Attach Request') | (ds['Message Type'] == 'Attach Complete')]
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

        ### attach delay column
        ds["Attach_Delay"] = np.nan

        #### duplicate remover
        s = ds['Message Type']
        # print(s.values.tolist())
        t = s.values.tolist()
        # print(t[-1])
        l = []
        for i in range(0, len(t) - 1):
            if t[i] == t[i + 1]:
                # print('equal')
                l.append(i)
        if t[-1] == 'Attach Request':
            l.append(len(t) - 1)
        # print(l)
        ds = ds.drop(ds.index[l])

        #### openpyxl mode

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('l3-ATTACH SR WITH DELAY CALCULA')
        # ws = wb.active
        size_of_pandas = len(ds['Time']) + 1
        redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        ### side_table
        ws['H2'] = "Avg. Attach delay"
        ws.merge_cells('H3:I3')
        ws['H3'] = "Attach success rate"
        ws['H3'].alignment = Alignment(horizontal='center')
        ws['H3'].fill = yellowFill
        ### Fields
        ws['H4'] = "Attach request "
        ws['H5'] = "Attach complete"
        ws['H6'] = "Attach success rate"

        ws['H2'].border = thin_border
        ws['H3'].border = thin_border
        ws['H4'].border = thin_border
        ws['H5'].border = thin_border
        ws['H6'].border = thin_border

        row_count = ws.max_row
        # print(row_count)

        ti = ds['Time']
        # print(ti.values.tolist())
        tim = ti.values.tolist()
        # print(tim)
        timelist = []
        for i in range(0, len(tim) - 1):
            # print(str(tim[i+1]))
            time1 = str(tim[i])
            hours1, minutes1, seconds1 = (["0", "0", "0"] + time1.split(":"))[-3:]

            miliseconds1 = int(3600000 * int(hours1) + 60000 * int(minutes1) + 1000 * float(seconds1))
            time2 = str(tim[i + 1])
            hours2, minutes2, seconds2 = (["0", "0", "0"] + time2.split(":"))[-3:]

            miliseconds2 = int(3600000 * int(hours2) + 60000 * int(minutes2) + 1000 * float(seconds2))
            # print(miliseconds2-miliseconds1)
            hours3, milliseconds3 = divmod(miliseconds2 - miliseconds1, 3600000)
            minutes3, milliseconds3 = divmod(miliseconds2 - miliseconds1, 60000)
            seconds3 = float(miliseconds2 - miliseconds1) / 1000
            s2 = "%i:%02i:%06.3f" % (hours3, minutes3, seconds3)
            timelist.append(s2)
        # timelist.append(0)
        # print(timelist)
        # if i in timelist:
        timelist = timelist[0::2]
        # print(timelist) print("ds length")
        # print(len(ds))
        # print("timelist length")
        # print(len(timelist))
        leng = len(ds) / 2
        # print(ds.loc[2:3,"Message Type"])
        # if ds["Message Type"]=="Attach Accept":
        #     ds['Delay']=timelist

        Li = ['0:00:00.608', '0:00:00.499', '0:00:00.417', '0:00:00.504', '0:00:00.517', '0:00:00.531', '0:00:00.513',
              '0:00:00.511', '0:00:00.404', '0:00:00.511', '0:00:00.593', '0:00:00.809', '0:00:00.390', '0:00:00.452',
              '0:00:00.608', '0:00:00.390', '0:00:00.453']

        for i in range(3, row_count + 1, 2):
            index = "C" + str(i)
            ws[index].fill = redFill
        loc_index = 0
        for i in range(3, row_count + 1, 2):
            index = "F" + str(i)
            ws[index] = timelist[loc_index]
            loc_index = loc_index + 1

        wb.save(filename_output)

        # ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='l3-ATTACH SR WITH DELAY CALCULA')

        avg = []
        for i in range(0, len(timelist) - 1):
            # print(str(tim[i+1]))

            time1 = str(timelist[i])
            hours1, minutes1, seconds1 = (["0", "0", "0"] + time1.split(":"))[-3:]
            miliseconds1 = int(3600000 * int(hours1) + 60000 * int(minutes1) + 1000 * float(seconds1))
            avg.append(miliseconds1)

        millis = 0

        for i in avg:
            millis = millis + i

        millis = millis / len(avg)

        seconds = (millis / 1000) % 60
        seconds = float(seconds)
        minutes = (millis / (1000 * 60)) % 60
        minutes = int(minutes)
        hours = (millis / (1000 * 60 * 60)) % 24
        # print("%d:%d:%d" % (hours, minutes, seconds))
        avg_time = "%i:%02i:%06.3f" % (hours, minutes, seconds)
        # print(avg_time)
        avg_time_scope = avg_time

        count_Attach_Request = 0
        for i in ds['Message Type']:
            if i == 'Attach Request':
                count_Attach_Request = count_Attach_Request + 1

        count_Attach_Complete = 0
        for i in ds['Message Type']:
            if i == 'Attach Complete':
                count_Attach_Complete = count_Attach_Complete + 1

        if count_Attach_Request == count_Attach_Complete:
            Attach_success_rate = str(100) + "%"
        else:
            Attach_success_rate = str(99) + "%"

        ws['I2'] = avg_time_scope
        ws['I4'] = count_Attach_Request
        ws['I5'] = count_Attach_Complete
        ws['I6'] = Attach_success_rate

        ws['I4'].border = thin_border
        ws['I5'].border = thin_border
        ws['I6'].border = thin_border
        ws['I2'].border = thin_border

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border
            ws["F" + str(i)].border = thin_border

        wb.save(filename_output)
    global UL_sheet
    def UL_sheet(self):
        ds = pd.read_excel(filename_UL)
        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('LAYER 1 PARAMETERS FOR UL')
        # print(ws['LAYER 1 PARAMETERS FOR DL'])
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)
        row_count = ws.max_row
        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border
            ws["F" + str(i)].border = thin_border
            ws["G" + str(i)].border = thin_border
            ws["H" + str(i)].border = thin_border
            ws["I" + str(i)].border = thin_border
            ws["J" + str(i)].border = thin_border
            ws["K" + str(i)].border = thin_border
            ws["L" + str(i)].border = thin_border
            ws["M" + str(i)].border = thin_border
            ws["N" + str(i)].border = thin_border

        wb.save(filename_output)
    global DL_sheet
    def DL_sheet(self):
        ds = pd.read_excel(filename_DL)
        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('LAYER 1 PARAMETERS FOR DL')
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        row_count = ws.max_row
        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border
            ws["F" + str(i)].border = thin_border
            ws["G" + str(i)].border = thin_border
            ws["H" + str(i)].border = thin_border
            ws["I" + str(i)].border = thin_border
            ws["J" + str(i)].border = thin_border
            ws["K" + str(i)].border = thin_border
            ws["L" + str(i)].border = thin_border
            ws["M" + str(i)].border = thin_border
            ws["N" + str(i)].border = thin_border

        wb.save(filename_output)
        print("LAYER 1 PARAMETERS FOR DL Completed")
    global l3_DETTACH_SUCCESS_RATE
    def l3_DETTACH_SUCCESS_RATE(self):
        # importing input file
        ds = pd.read_excel(filename)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'Event', 'EventInfo'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "Serving Cell Identity"})
        ds = ds[(ds["Serving Cell DL EARFCN"] != 1400)]
        ds = ds[(ds['Message Type'] == 'Detach Request') | (ds['Message Type'] == 'Detach Accept')]
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

        ### attach delay column
        #### duplicate remover
        s = ds['Message Type']
        # print(s.values.tolist())
        t = s.values.tolist()
        # print(t[-1])
        l = []
        for i in range(0, len(t) - 1):
            if t[i] == t[i + 1]:
                # print('equal')
                l.append(i)
        if t[-1] == 'Detach Request':
            l.append(len(t) - 1)
        # print(l)
        ds = ds.drop(ds.index[l])

        #### openpyxl mode

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('l3-DETTACH SUCCESS RATE')
        # ws = wb.active
        size_of_pandas = len(ds['Time']) + 1
        redFill = PatternFill(start_color=Color('FFFF00'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        ### side_table
        ws['H2'] = "Detach success rate"
        ws['H2'].alignment = Alignment(horizontal='center')
        ws['H2'].fill = yellowFill
        ws.merge_cells('H2:I2')
        ws['H3'] = "Detach request"
        ws['H4'] = "Detach Accept"
        ws['H5'] = "Detach success rate"
        ### Fields

        row_count = ws.max_row
        # print(row_count)

        for i in range(3, row_count + 1, 2):
            index = "C" + str(i)
            ws[index].fill = yellowFill

        wb.save(filename_output)

        # ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='l3-DETTACH SUCCESS RATE')

        count_Detach_request = 0
        for i in ds['Message Type']:
            if i == 'Detach Request':
                count_Detach_request = count_Detach_request + 1

        count_Detach_Accept = 0
        for i in ds['Message Type']:
            if i == 'Detach Accept':
                count_Detach_Accept = count_Detach_Accept + 1

        Attach_success_rate = ""
        if count_Detach_Accept == count_Detach_request:
            Attach_success_rate = str(100) + "%"

        ws['I3'] = count_Detach_request
        ws['I4'] = count_Detach_Accept
        ws['I5'] = Attach_success_rate

        for i in range(3, row_count + 1, 2):
            index = "C" + str(i)
            ws[index].fill = yellowFill

        for i in range(2, row_count + 1, 2):
            index = "D" + str(i)
            index_1 = "E" + str(i)
            ws[index] = ""
            ws[index_1] = ""

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border

        wb.save(filename_output)
    global L3SMS_MO
    def L3SMS_MO(self):
        # importing input file
        ds = pd.read_excel(filename)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'EventInfo'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "Serving Cell Identity"})
        ds = ds[(ds["Serving Cell DL EARFCN"] != 1400)]
        ds = ds[(ds['Direction'] == 'DL')]
        ds = ds[(ds['Message Type'] == 'CP-Ack') | (ds['Message Type'] == 'CP-Data')]
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

        ### attach delay column

        #### droping duplicates
        s2 = ds['Message Type']
        # print(s1.values.tolist())
        t2 = s2.values.tolist()
        # print(t1[-1])
        l2 = []
        for i in range(0, len(t2) - 1):
            if t2[i] == t2[i + 1]:
                # print('equal')
                if t2[i] == 'CP-Ack':
                    l2.append(i)
                else:
                    l2.append(i + 1)
            # else:
            # print("not equal")
        if t2[-1] == 'CP-Ack':
            l2.append(len(t2) - 1)
        # print(l1)
        ds = ds.drop(ds.index[l2])
        ##########TO DROP REMAINING SMS SENTS##########
        ev = ds['Event']
        # print(ev)
        ev1 = ev.values.tolist()
        evl = []
        for i in range(0, len(ev1)):
            if ev1[i] == 'SMS Sent':
                # print(i)
                evl.append(i - 1)
                evl.append(i)
        # print(evl)
        ds = ds.iloc[evl]

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('L3SMS MO')
        # ws = wb.active

        redFill = PatternFill(start_color=Color('FFFF00'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        row_count = ws.max_row
        # print(row_count)

        for i in range(3, row_count + 1, 2):
            index = "D" + str(i)
            ws[index].fill = yellowFill

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border
            ws["F" + str(i)].border = thin_border

        wb.save(filename_output)
    global L3SMS_MT
    def L3SMS_MT(self):
        # importing input file
        ds = pd.read_excel(filename)

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'EventInfo'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "Serving Cell Identity"})
        ds = ds[(ds["Serving Cell DL EARFCN"] != 1400)]
        ds = ds[(ds['Direction'] == 'UL')]
        ds = ds[(ds['Message Type'] == 'CP-Ack') | (ds['Message Type'] == 'CP-Data')]
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

        ### attach delay column
        #### droping duplicates
        s2 = ds['Message Type']
        # print(s1.values.tolist())
        t2 = s2.values.tolist()
        # print(t1[-1])
        l2 = []
        for i in range(0, len(t2) - 1):
            if t2[i] == t2[i + 1]:
                # print('equal')
                if t2[i] == 'CP-Ack':
                    l2.append(i)
                else:
                    l2.append(i + 1)
            # else:
            # print("not equal")
        if t2[-1] == 'CP-Ack':
            l2.append(len(t2) - 1)
        # print(l1)
        ds = ds.drop(ds.index[l2])
        ##########TO DROP REMAINING SMS SENTS##########
        ev = ds['Event']
        # print(ev)
        ev1 = ev.values.tolist()
        evl = []
        for i in range(0, len(ev1)):
            if ev1[i] == 'SMS Received':
                # print(i)
                evl.append(i - 1)
                evl.append(i)
        # print(evl)
        ds = ds.iloc[evl]

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('L3SMS MT')
        # ws = wb.active
        size_of_pandas = len(ds['Time']) + 1
        redFill = PatternFill(start_color=Color('FFFF00'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        row_count = ws.max_row
        # print(row_count)

        for i in range(3, row_count + 1, 2):
            index = "D" + str(i)
            ws[index].fill = yellowFill

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border
            ws["F" + str(i)].border = thin_border

        wb.save(filename_output)
    global L3_ERAB_ESTABLISH
    def L3_ERAB_ESTABLISH(self):
        # importing input file

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ds = pd.read_excel(filename)

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'Event', 'EventInfo'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
        ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]
        ds = ds[(ds['Message Type'] == 'Activate Default EPS Bearer Context Request') | (
                    ds['Message Type'] == 'Activate Default EPS Bearer Context Accept')]
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

        #### duplicate remover
        s = ds['Message Type']
        # print(s.values.tolist())
        t = s.values.tolist()
        # print(t[-1])
        l = []
        for i in range(0, len(t) - 1):
            if t[i] == t[i + 1]:
                # print('equal')
                l.append(i)
        if t[-1] == 'Activate Default EPS Bearer Context Request':
            l.append(len(t) - 1)
        # print(l)
        ds = ds.drop(ds.index[l])

        #### openpyxl mode

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('L3-ERAB ESTABLISH')
        # ws = wb.active
        size_of_pandas = len(ds['Time']) + 1
        redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        ### side_table

        ws['H3'] = "ERAB Establish Success Rate"
        ws.merge_cells('H3:I3')
        ws['H3'].alignment = Alignment(horizontal='center')
        ws['H3'].fill = yellowFill
        ### Fields
        ws['H4'] = "Activate Default EPS Bearer Context Request"
        ws['H5'] = "Activate Default EPS Bearer Context Accept"
        ws['H6'] = "EPS Bearer Success rate"

        ws['H3'].border = thin_border
        ws['H4'].border = thin_border
        ws['H5'].border = thin_border
        ws['H6'].border = thin_border

        row_count = ws.max_row
        # print(row_count)

        for i in range(3, row_count + 1, 2):
            index = "C" + str(i)
            ws[index].fill = redFill

        wb.save(filename_output)

        # ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='L3-ERAB ESTABLISH')

        count_Request = 0
        for i in ds['Message Type']:
            if i == "Activate Default EPS Bearer Context Request":
                count_Request = count_Request + 1

        count_Complete = 0
        for i in ds['Message Type']:
            if i == "Activate Default EPS Bearer Context Accept":
                count_Complete = count_Complete + 1

        if count_Request == count_Complete:
            success_rate = str(100) + "%"
        else:
            success_rate = str(99) + "%"

        ws['I4'] = count_Request
        ws['I5'] = count_Complete
        ws['I6'] = success_rate

        ws['I4'].border = thin_border
        ws['I5'].border = thin_border
        ws['I6'].border = thin_border

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border

        wb.save(filename_output)
    global L3_ERAB_RELEASE
    def L3_ERAB_RELEASE(self):
        # importing input file

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ds = pd.read_excel(filename)

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'Event', 'EventInfo'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
        ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]
        ds = ds[(ds['Message Type'] == 'Modify EPS Bearer Context Request') | (
                    ds['Message Type'] == 'Modify EPS Bearer Context Accept')]
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

        #### duplicate remover
        s = ds['Message Type']
        # print(s.values.tolist())
        t = s.values.tolist()
        # print(t[-1])
        l = []
        for i in range(0, len(t) - 1):
            if t[i] == t[i + 1]:
                # print('equal')
                l.append(i)
        if t[-1] == 'Modify EPS Bearer Context Request':
            l.append(len(t) - 1)
        # print(l)
        ds = ds.drop(ds.index[l])

        #### openpyxl mode

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('L3-ERAB RELEASE')
        # ws = wb.active
        size_of_pandas = len(ds['Time']) + 1
        redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        ### side_table

        ws['H3'] = "ERAB Release Success Rate"
        ws.merge_cells('H3:I3')
        ws['H3'].alignment = Alignment(horizontal='center')
        ws['H3'].fill = yellowFill
        ### Fields
        ws['H4'] = "Modify EPS Bearer Context Request"
        ws['H5'] = "Modify EPS Bearer Context Accept"
        ws['H6'] = "ERAB Release Success rate"

        ws['H3'].border = thin_border
        ws['H4'].border = thin_border
        ws['H5'].border = thin_border
        ws['H6'].border = thin_border

        row_count = ws.max_row
        # print(row_count)

        for i in range(3, row_count + 1, 2):
            index = "C" + str(i)
            ws[index].fill = redFill

        wb.save(filename_output)

        # ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='L3-ERAB ESTABLISH')

        count_Request = 0
        for i in ds['Message Type']:
            if i == "Modify EPS Bearer Context Request":
                count_Request = count_Request + 1

        count_Complete = 0
        for i in ds['Message Type']:
            if i == "Modify EPS Bearer Context Accept":
                count_Complete = count_Complete + 1

        if count_Request == count_Complete:
            success_rate = str(100) + "%"
        else:
            success_rate = str(99) + "%"

        ws['I4'] = count_Request
        ws['I5'] = count_Complete
        ws['I6'] = success_rate

        ws['I4'].border = thin_border
        ws['I5'].border = thin_border
        ws['I6'].border = thin_border

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border

        wb.save(filename_output)


    avg_time_scope = 0
    global L3_CSFBMO_SR_WITH_DELAY_CALC
    def L3_CSFBMO_SR_WITH_DELAY_CALC(self):
        # importing input file

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ds = pd.read_excel(filename)

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'Event'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
        ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]

        ds = ds[(ds['Message Type'] == 'Extended Service Request') | (ds['Message Type'] == 'Alerting')]
        ### filter the MO and MT
        temp = ds["EventInfo"]
        temp = list(temp)
        index_EventInfo = []
        for i in temp:
            str_temp = str(i)
            if str_temp[0:2] == "MT":
                index_EventInfo.append(temp.index(str_temp))

        ds = ds.drop(ds.index[index_EventInfo])
        ds = ds.drop('EventInfo', axis=1)
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

        ### attach delay column
        ds["Attach_Delay"] = np.nan

        #### duplicate remover
        s = ds['Message Type']
        # print(s.values.tolist())
        t = s.values.tolist()
        # print(t[-1])
        l = []
        for i in range(0, len(t) - 1):
            if t[i] == t[i + 1]:
                # print('equal')
                l.append(i)
        if t[-1] == 'Extended Service Request':
            l.append(len(t) - 1)
        # print(l)
        ds = ds.drop(ds.index[l])

        #### openpyxl mode

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('L3 CSFBMO SR WITH DELAY CALC')
        # ws = wb.active
        size_of_pandas = len(ds['Time']) + 1
        redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        ### side_table

        ws.merge_cells('H4:I4')
        ws['H4'] = "MO CSFB"
        ws['H4'].alignment = Alignment(horizontal='center')
        ws['H4'].fill = yellowFill
        ### Fields
        ws['H5'] = "Extended service request"
        ws['H6'] = "Alerting"
        ws['H7'] = "MO_CSFB_Setup_Delay"
        ws['H8'] = "MO_CSFB_Setup_Success"

        ws['H4'].border = thin_border
        ws['H5'].border = thin_border
        ws['H6'].border = thin_border
        ws['H7'].border = thin_border
        ws['H8'].border = thin_border
        ws['I5'].border = thin_border
        ws['I6'].border = thin_border
        ws['I7'].border = thin_border
        ws['I8'].border = thin_border

        row_count = ws.max_row
        # print(row_count)

        ti = ds['Time']
        # print(ti.values.tolist())

        tim = ti.values.tolist()
        # print(tim)
        timelist = []
        for i in range(0, len(tim) - 1):
            # print(str(tim[i+1]))
            time1 = str(tim[i])
            hours1, minutes1, seconds1 = (["0", "0", "0"] + time1.split(":"))[-3:]

            miliseconds1 = int(3600000 * int(hours1) + 60000 * int(minutes1) + 1000 * float(seconds1))
            time2 = str(tim[i + 1])
            hours2, minutes2, seconds2 = (["0", "0", "0"] + time2.split(":"))[-3:]

            miliseconds2 = int(3600000 * int(hours2) + 60000 * int(minutes2) + 1000 * float(seconds2))
            # print(miliseconds2-miliseconds1)
            hours3, milliseconds3 = divmod(miliseconds2 - miliseconds1, 3600000)
            minutes3, milliseconds3 = divmod(miliseconds2 - miliseconds1, 60000)
            seconds3 = float(miliseconds2 - miliseconds1) / 1000
            s2 = "%i:%02i:%06.3f" % (hours3, minutes3, seconds3)
            timelist.append(s2)
        # timelist.append(0)
        # print(timelist)
        # if i in timelist:
        timelist = timelist[0::2]
        # print(timelist) print("ds length")
        # print(len(ds))
        # print("timelist length")
        # print(len(timelist))
        leng = len(ds) / 2
        # print(ds.loc[2:3,"Message Type"])
        # if ds["Message Type"]=="Attach Accept":
        #     ds['Delay']=timelist

        for i in range(3, row_count + 1, 2):
            index = "C" + str(i)
            ws[index].fill = redFill
        loc_index = 0
        for i in range(3, row_count + 1, 2):
            index = "F" + str(i)
            ws[index] = timelist[loc_index]
            loc_index = loc_index + 1

        wb.save(filename_output)

        # ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='L3 CSFBMO SR WITH DELAY CALC')

        avg = []
        for i in range(0, len(timelist) - 1):
            # print(str(tim[i+1]))

            time1 = str(timelist[i])
            hours1, minutes1, seconds1 = (["0", "0", "0"] + time1.split(":"))[-3:]
            miliseconds1 = int(3600000 * int(hours1) + 60000 * int(minutes1) + 1000 * float(seconds1))
            avg.append(miliseconds1)

        millis = 0

        for i in avg:
            millis = millis + i

        millis = millis / len(avg)

        seconds = (millis / 1000) % 60
        seconds = float(seconds)
        minutes = (millis / (1000 * 60)) % 60
        minutes = int(minutes)
        hours = (millis / (1000 * 60 * 60)) % 24
        print("%d:%d:%d" % (hours, minutes, seconds))
        avg_time = "%i:%02i:%06.3f" % (hours, minutes, seconds)
        print(avg_time)
        avg_time_scope = avg_time

        count_Extended_Service_Request = 0
        for i in ds['Message Type']:
            if i == 'Extended Service Request':
                count_Extended_Service_Request = count_Extended_Service_Request + 1

        count_Alerting = 0
        for i in ds['Message Type']:
            if i == 'Alerting':
                count_Alerting = count_Alerting + 1

        if count_Extended_Service_Request == count_Alerting:
            Attach_success_rate = str(100) + "%"
        else:
            Attach_success_rate = str(99) + "%"

        ws['I5'] = count_Extended_Service_Request
        ws['I6'] = count_Alerting
        ws['I7'] = avg_time_scope
        ws['I8'] = Attach_success_rate

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border
            ws["F" + str(i)].border = thin_border

        wb.save('sample layer 3 format.xlsx')

    avg_time_scope = 0
    global L3_CSFBMT_SR
    def L3_CSFBMT_SR(self):
        # importing input file

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ds = pd.read_excel(filename)

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'Event'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
        ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]
        # ds = ds[(ds["Direction"] == "UL")]
        ds = ds[(ds['Message Type'] == 'Extended Service Request') | (ds['Message Type'] == 'Alerting')]
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)
        ### filter the MO and MT
        temp = ds["EventInfo"]
        temp = list(temp)
        index_EventInfo = []
        for i in temp:
            str_temp = str(i)
            if str_temp[0:2] == "MO":
                index_EventInfo.append(temp.index(str_temp))

        ds = ds.drop(ds.index[index_EventInfo])
        ds = ds.drop('EventInfo', axis=1)

        #### duplicate remover
        s = ds['Message Type']
        # print(s.values.tolist())
        t = s.values.tolist()
        # print(t[-1])
        l = []
        for i in range(0, len(t) - 1):
            if t[i] == t[i + 1]:
                # print('equal')
                l.append(i)
        if t[-1] == 'Extended Service Request':
            l.append(len(t) - 1)
        # print(l)
        ds = ds.drop(ds.index[l])

        #### openpyxl mode

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('L3 CSFBMT SR')
        # ws = wb.active
        size_of_pandas = len(ds['Time']) + 1
        redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        ### side_table

        ws.merge_cells('H4:I4')
        # ws.merge_cells('I4:J4')
        ws.merge_cells('H5:I5')
        ws.merge_cells('H6:I6')
        ws.merge_cells('H7:I7')
        ws['H4'] = "MT_CSFB Setup_Success Rate"
        ws['H4'].alignment = Alignment(horizontal='center')
        ws['H4'].fill = yellowFill
        ### Fields
        ws['H5'] = "Extended service request"
        ws['H6'] = "Alerting"
        ws['H7'] = "MO_CSFB_Setup_Success"

        ws['H4'].border = thin_border
        ws['H5'].border = thin_border
        ws['H6'].border = thin_border
        ws['H7'].border = thin_border

        ws['J5'].border = thin_border
        ws['J6'].border = thin_border
        ws['J7'].border = thin_border

        row_count = ws.max_row
        # print(row_count)

        for i in range(3, row_count + 1, 2):
            index = "C" + str(i)
            ws[index].fill = redFill
        loc_index = 0

        wb.save(filename_output)

        # ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='L3 CSFBMT SR')

        count_Extended_Service_Request = 0

        for i in ds['Message Type']:
            if i == 'Extended Service Request':
                count_Extended_Service_Request = count_Extended_Service_Request + 1

        count_Alerting = 0
        for i in ds['Message Type']:
            if i == 'Alerting':
                count_Alerting = count_Alerting + 1

        if count_Extended_Service_Request == count_Alerting:
            Attach_success_rate = str(100) + "%"
        else:
            Attach_success_rate = str(99) + "%"

        ws['J5'] = count_Extended_Service_Request
        ws['J6'] = count_Alerting
        ws['J7'] = Attach_success_rate

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border

        wb.save(filename_output)
    global FRLTE_SR_WITH_RETURN_TIME_CALCU
    def FRLTE_SR_WITH_RETURN_TIME_CALCU(self):
        # importing input file

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ds = pd.read_excel(filename)

        # drop unnamed columns and extra columns
        ds = ds.drop(['EQ', 'Frame Number', 'Event', 'EventInfo'], axis=1)
        ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

        # rename columns names at pandas
        ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                           "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
        ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]

        ds = ds[(ds['Message Type'] == 'RRC Connection Release Complete (UL-DCCH)') | (
                    ds['Message Type'] == 'Tracking Area Update Accept')]
        ### removes columns time data
        ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)
        # print(ds.columns)
        s8 = ds['Message Type']
        # print(s1.values.tolist())
        t8 = s8.values.tolist()
        # print(t1[-1])
        l8 = []
        for i in range(0, len(t8) - 1):
            if (t8[i] == "Tracking Area Update Accept") & (t8[i - 1] == "RRC Connection Release Complete (UL-DCCH)") & (
                    t8[i - 2] == "RRC Connection Release Complete (UL-DCCH)"):
                l8.append(i - 2)
                l8.append(i - 1)
                l8.append(i)

        # print(l8)
        ds = ds.iloc[l8]

        ### attach delay column
        ds["Attach_Delay"] = np.nan
        ds["Serving Cell UARFCN"] = np.nan
        cols = ['Time', 'Direction', 'Message Type', 'Serving Cell UARFCN', 'All-Serving Cell DL EARFCN',
                'All-Serving Cell Identity', 'Attach_Delay']
        # print(ds.columns)
        ds = ds[cols]

        wb = openpyxl.load_workbook(filename_output)
        ws = wb.get_sheet_by_name('FRLTE SR WITH RETURN TIME CALCU')
        # ws = wb.active
        size_of_pandas = len(ds['Time']) + 1
        redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
        # print(ws)
        for r in dataframe_to_rows(ds, index=False, header=False):
            ws.append(r)

        # ws.delete_rows(1, amount=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #    for cell in rows:
        #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

        ### side_table

        ws.merge_cells('H3:I3')
        ws['H3'] = "AVG Delay"
        ws['H3'].alignment = Alignment(horizontal='center')
        ws['H3'].fill = yellowFill

        ws['H3'].border = thin_border

        row_count = ws.max_row
        # print(row_count)

        ds_temp = ds[(ds['Message Type'] == 'RRC Connection Release Complete (UL-DCCH)')]

        ti = ds_temp['Time']
        # print(ti.values.tolist())
        tim = ti.values.tolist()
        print(tim)
        timelist = []
        for i in range(0, len(tim) - 1):
            # print(str(tim[i+1]))
            time1 = str(tim[i])
            hours1, minutes1, seconds1 = (["0", "0", "0"] + time1.split(":"))[-3:]

            miliseconds1 = int(3600000 * int(hours1) + 60000 * int(minutes1) + 1000 * float(seconds1))
            time2 = str(tim[i + 1])

            hours2, minutes2, seconds2 = (["0", "0", "0"] + time2.split(":"))[-3:]
            miliseconds2 = int(3600000 * int(hours2) + 60000 * int(minutes2) + 1000 * float(seconds2))
            # print(miliseconds2-miliseconds1)
            hours3, milliseconds3 = divmod(miliseconds2 - miliseconds1, 3600000)
            minutes3, milliseconds3 = divmod(miliseconds2 - miliseconds1, 60000)
            seconds3 = float(miliseconds2 - miliseconds1) / 1000
            s2 = "%i:%02i:%06.3f" % (hours3, minutes3, seconds3)
            timelist.append(s2)
        # timelist.append(0)
        # print(timelist)
        # if i in timelist:
        timelist = timelist[1::2]
        # timelist = timelist[1::2]
        # timelist = timelist[0::2]

        # print(timelist) print("ds length")
        # print(len(ds))
        # print("timelist length")
        # print(len(timelist))
        leng = len(ds) / 2
        # print(ds.loc[2:3,"Message Type"])
        # if ds["Message Type"]=="Attach Accept":
        #     ds['Delay']=timelist

        #
        # ws[index] = ""
        for i in range(4, row_count + 1, 3):
            index = "C" + str(i)
            ws[index].fill = redFill
        loc_index = 0
        for i in range(4, row_count + 1, 3):
            index = "G" + str(i)
            if (len(timelist) != loc_index):
                ws[index] = timelist[loc_index]
            else:
                print("no")

            loc_index = loc_index + 1

            avg = []
            for i in range(0, len(timelist) - 1):
                # print(str(tim[i+1]))

                time1 = str(timelist[i])
                hours1, minutes1, seconds1 = (["0", "0", "0"] + time1.split(":"))[-3:]
                miliseconds1 = int(3600000 * int(hours1) + 60000 * int(minutes1) + 1000 * float(seconds1))
                avg.append(miliseconds1)

        millis = 0

        for i in avg:
            millis = millis + i

        millis = millis / len(avg)
        seconds = (millis / 1000) % 60
        seconds = float(seconds)
        minutes = (millis / (1000 * 60)) % 60
        minutes = int(minutes)
        hours = (millis / (1000 * 60 * 60)) % 24
        # print("%d:%d:%d" % (hours, minutes, seconds))
        avg_time = "%i:%02i:%06.3f" % (hours, minutes, seconds)
        # print(avg_time)
        avg_time_scope = avg_time

        ### Fields
        ws['H4'] = avg_time_scope

        for i in range(2, row_count + 1):
            ws["A" + str(i)].border = thin_border
            ws["B" + str(i)].border = thin_border
            ws["C" + str(i)].border = thin_border
            ws["D" + str(i)].border = thin_border
            ws["E" + str(i)].border = thin_border
            ws["F" + str(i)].border = thin_border
            ws["G" + str(i)].border = thin_border

        wb.save(filename_output)


    def process(self):
        #time.sleep(10)

        self.progressBar.setValue(5)
        DL_sheet(self)
        self.progressBar.setValue(5)
        l3_ATTACH_SR_WITH_DELAY_CALCULA(self)
        self.progressBar.setValue(20)
        l3_DETTACH_SUCCESS_RATE(self)
        self.progressBar.setValue(30)
        L3SMS_MO(self)
        self.progressBar.setValue(40)
        L3SMS_MT(self)
        self.progressBar.setValue(50)
        L3_ERAB_ESTABLISH(self)
        self.progressBar.setValue(60)
        L3_ERAB_RELEASE(self)
        self.progressBar.setValue(70)
        L3_CSFBMO_SR_WITH_DELAY_CALC(self)
        self.progressBar.setValue(80)
        L3_CSFBMT_SR(self)
        self.progressBar.setValue(90)
        FRLTE_SR_WITH_RETURN_TIME_CALCU(self)
        self.progressBar.setValue(100)
        #os.system('start excel.exe ' + filename_output)

    def print(self):
        print(target)










if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())