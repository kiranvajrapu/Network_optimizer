import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Side,Border
thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
def UL_sheet():
    ds = pd.read_excel("UL MADHAVARAM.xlsx")
    wb = openpyxl.load_workbook('sample layer 3 format.xlsx')
    ws = wb.get_sheet_by_name('LAYER 1 PARAMETERS FOR UL')
    #print(ws['LAYER 1 PARAMETERS FOR DL'])
    for r in dataframe_to_rows(ds, index=False, header=False):
        ws.append(r)
    row_count = ws.max_row
    for i in range(2, row_count + 1):
        ws["A" + str(i)].border = thin_border
        ws["B" + str(i)].border = thin_border
        ws["C" + str(i)].border = thin_border
        ws["D" + str(i)].border = thin_border
        ws["E" + str(i)].border = thin_border
        ws["F" + str(i)].border = thin_border
        ws["G" + str(i)].border = thin_border
        ws["H" + str(i)].border = thin_border
        ws["I" + str(i)].border = thin_border
        ws["J" + str(i)].border = thin_border
        ws["K" + str(i)].border = thin_border
        ws["L" + str(i)].border = thin_border
        ws["M" + str(i)].border = thin_border
        ws["N" + str(i)].border = thin_border

    wb.save('sample layer 3 format.xlsx')

