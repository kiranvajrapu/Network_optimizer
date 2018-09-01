import pandas as pd
import numpy as np
import csv
import os
import xlrd
import sys
import datetime
import Dataset
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Color,fills,Side,Border,GradientFill,Font,Alignment


def l3_DETTACH_SUCCESS_RATE():
    # importing input file
    ds = Dataset.dataset_extract()
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    #drop unnamed columns and extra columns
    ds = ds.drop(['EQ', 'Frame Number', 'Event', 'EventInfo'], axis=1)
    ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

    #rename columns names at pandas
    ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "Serving Cell DL EARFCN",
                                       "All-Serving Cell Identity[1]": "Serving Cell Identity"})
    ds = ds[(ds["Serving Cell DL EARFCN"] != 1400)]
    ds = ds[(ds['Message Type'] == 'Detach Request') | (ds['Message Type'] == 'Detach Accept')]
    ### removes columns time data
    ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

    ### attach delay column
    #### duplicate remover
    s = ds['Message Type']
    # print(s.values.tolist())
    t = s.values.tolist()
    # print(t[-1])
    l = []
    for i in range(0, len(t) - 1):
        if t[i] == t[i + 1]:
            # print('equal')
            l.append(i)
    if t[-1] == 'Detach Request':
        l.append(len(t) - 1)
    # print(l)
    ds = ds.drop(ds.index[l])

    #### openpyxl mode

    wb = openpyxl.load_workbook('sample layer 3 format.xlsx')
    ws = wb.get_sheet_by_name('l3-DETTACH SUCCESS RATE')
    #ws = wb.active
    size_of_pandas = len(ds['Time']) + 1
    redFill = PatternFill(start_color=Color('FFFF00'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    #print(ws)
    for r in dataframe_to_rows(ds, index=False, header=False):
        ws.append(r)

    # ws.delete_rows(1, amount=1)

    # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
    #    for cell in rows:
    #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

    ### side_table
    ws['H2'] = "Detach success rate"
    ws['H2'].alignment = Alignment(horizontal='center')
    ws['H2'].fill = yellowFill
    ws.merge_cells('H2:I2')
    ws['H3'] = "Detach request"
    ws['H4'] = "Detach Accept"
    ws['H5'] = "Detach success rate"
    ### Fields




    row_count = ws.max_row
    #print(row_count)


    for i in range(3, row_count + 1, 2):
        index = "C" + str(i)
        ws[index].fill = yellowFill


    wb.save('sample layer 3 format.xlsx')

    #ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='l3-DETTACH SUCCESS RATE')

    count_Detach_request = 0
    for i in ds['Message Type']:
        if i == 'Detach Request':
            count_Detach_request = count_Detach_request + 1

    count_Detach_Accept = 0
    for i in ds['Message Type']:
        if i == 'Detach Accept':
            count_Detach_Accept = count_Detach_Accept + 1

    Attach_success_rate = ""
    if count_Detach_Accept==count_Detach_request:
        Attach_success_rate = str(100) + "%"

    ws['I3'] = count_Detach_request
    ws['I4'] = count_Detach_Accept
    ws['I5'] = Attach_success_rate

    for i in range(3, row_count + 1, 2):
        index = "C" + str(i)
        ws[index].fill = yellowFill

    for i in range(2,row_count + 1,2):
        index = "D" + str(i)
        index_1 = "E" + str(i)
        ws[index] = ""
        ws[index_1] = ""

    for i in range(2, row_count + 1):
        ws["A" + str(i)].border = thin_border
        ws["B" + str(i)].border = thin_border
        ws["C" + str(i)].border = thin_border
        ws["D" + str(i)].border = thin_border
        ws["E" + str(i)].border = thin_border



    wb.save('sample layer 3 format.xlsx')






