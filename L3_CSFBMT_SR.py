import pandas as pd
import Dataset
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Color,fills,Side,Border,GradientFill,Font,Alignment
avg_time_scope = 0
def L3_CSFBMT_SR():
    # importing input file

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ds = Dataset.dataset_extract()

    #drop unnamed columns and extra columns
    ds = ds.drop(['EQ', 'Frame Number', 'Event'], axis=1)
    ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

    #rename columns names at pandas
    ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                       "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
    ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]
    #ds = ds[(ds["Direction"] == "UL")]
    ds = ds[(ds['Message Type'] == 'Extended Service Request') | (ds['Message Type'] == 'Alerting')]
    ### removes columns time data
    ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)
    ### filter the MO and MT
    temp = ds["EventInfo"]
    temp = list(temp)
    index_EventInfo = []
    for i in temp:
        str_temp = str(i)
        if str_temp[0:2] == "MO":
            index_EventInfo.append(temp.index(str_temp))

    ds = ds.drop(ds.index[index_EventInfo])
    ds = ds.drop('EventInfo', axis=1)

    #### duplicate remover
    s = ds['Message Type']
    #print(s.values.tolist())
    t = s.values.tolist()
    #print(t[-1])
    l = []
    for i in range(0, len(t) - 1):
        if t[i] == t[i + 1]:
            #print('equal')
            l.append(i)
    if t[-1] == 'Extended Service Request':
        l.append(len(t) - 1)
    # print(l)
    ds = ds.drop(ds.index[l])



    #### openpyxl mode



    wb = openpyxl.load_workbook('sample layer 3 format.xlsx')
    ws = wb.get_sheet_by_name('L3 CSFBMT SR')
    #ws = wb.active
    size_of_pandas = len(ds['Time']) + 1
    redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    #print(ws)
    for r in dataframe_to_rows(ds, index=False, header=False):
        ws.append(r)

    # ws.delete_rows(1, amount=1)

    # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
    #    for cell in rows:
    #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)

    ### side_table

    ws.merge_cells('H4:I4')
    #ws.merge_cells('I4:J4')
    ws.merge_cells('H5:I5')
    ws.merge_cells('H6:I6')
    ws.merge_cells('H7:I7')
    ws['H4'] = "MT_CSFB Setup_Success Rate"
    ws['H4'].alignment = Alignment(horizontal='center')
    ws['H4'].fill = yellowFill
    ### Fields
    ws['H5'] = "Extended service request"
    ws['H6'] = "Alerting"
    ws['H7'] = "MO_CSFB_Setup_Success"



    ws['H4'].border = thin_border
    ws['H5'].border = thin_border
    ws['H6'].border = thin_border
    ws['H7'].border = thin_border

    ws['J5'].border = thin_border
    ws['J6'].border = thin_border
    ws['J7'].border = thin_border


    row_count = ws.max_row
    #print(row_count)


    for i in range(3, row_count + 1, 2):
        index = "C" + str(i)
        ws[index].fill = redFill
    loc_index=0

    wb.save('sample layer 3 format.xlsx')

    #ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='L3 CSFBMT SR')




    count_Extended_Service_Request = 0

    for i in ds['Message Type']:
        if i == 'Extended Service Request':
            count_Extended_Service_Request = count_Extended_Service_Request + 1

    count_Alerting = 0
    for i in ds['Message Type']:
        if i == 'Alerting':
            count_Alerting = count_Alerting + 1

    if count_Extended_Service_Request==count_Alerting:
        Attach_success_rate = str(100) + "%"
    else:
        Attach_success_rate = str(99) + "%"

    ws['J5'] = count_Extended_Service_Request
    ws['J6'] = count_Alerting
    ws['J7'] = Attach_success_rate

    for i in range(2, row_count + 1):
        ws["A" + str(i)].border = thin_border
        ws["B" + str(i)].border = thin_border
        ws["C" + str(i)].border = thin_border
        ws["D" + str(i)].border = thin_border
        ws["E" + str(i)].border = thin_border


    wb.save('sample layer 3 format.xlsx')





