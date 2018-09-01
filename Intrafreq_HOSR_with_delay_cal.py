import pandas as pd
import numpy as np
import openpyxl
import Dataset
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Color,fills,Side,Border,GradientFill,Font,Alignment
avg_time_scope = 0
def Intrafreq_HOSR_with_delay_cal():
    # importing input file

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ds = pd.read_excel("Test_excel_sheet2.xlsx")

    #drop unnamed columns and extra columns
    ds = ds.drop(['EQ', 'Frame Number'], axis=1)
    ds = ds.loc[:, ~ds.columns.str.contains('^Unnamed')]

    #rename columns names at pandas
    ds = ds.rename(index=str, columns={"All-Serving Cell DL EARFCN[1]": "All-Serving Cell DL EARFCN",
                                       "All-Serving Cell Identity[1]": "All-Serving Cell Identity"})
    ds = ds[(ds["All-Serving Cell DL EARFCN"] != 1400)]




    ds = ds[(ds['Message Type'] == 'Measurement Report (UL-DCCH)') | (ds['Message Type'] == 'RRC Connection Reconfiguration Complete (UL-DCCH)')]



    ### removes columns time data
    ds['Time'] = ds['Time'].astype(str).str[:-3].astype(str)

    ### attach delay column
    ds["Attach_Delay"] = np.nan

    #### duplicate remover
    s = ds['Message Type']
    #print(s.values.tolist())
    t = s.values.tolist()
    #print(t[-1])
    l = []
    for i in range(0, len(t) - 1):
        if t[i] == t[i + 1]:
            #print('equal')
            l.append(i)
    if t[-1] == 'Extended Service Request':
        l.append(len(t) - 1)
    # print(l)
    ds = ds.drop(ds.index[l])



    #### openpyxl mode



    wb = openpyxl.load_workbook('sample layer 3 format.xlsx')
    ws = wb.get_sheet_by_name('Intrafreq  HOSR with delay cal')
    #ws = wb.active
    size_of_pandas = len(ds['Time']) + 1
    redFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    yellowFill = PatternFill(start_color=Color('008000'), end_color=Color('000000'), patternType=fills.FILL_SOLID)
    #print(ws)
    for r in dataframe_to_rows(ds, index=False, header=False):
        ws.append(r)

    # ws.delete_rows(1, amount=1)

    # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
    #    for cell in rows:
    #        cell.fill = PatternFill(start_color=Color('008000'),end_color=Color('000000'),patternType=fills.FILL_SOLID)


    row_count = ws.max_row
    #print(row_count)

    ti = ds['Time']
    # print(ti.values.tolist())

    tim = ti.values.tolist()
    #print(tim)
    timelist = []
    for i in range(0, len(tim) - 1):
        # print(str(tim[i+1]))
        time1 = str(tim[i])
        hours1, minutes1, seconds1 = (["0", "0", "0"] + time1.split(":"))[-3:]

        miliseconds1 = int(3600000 * int(hours1) + 60000 * int(minutes1) + 1000 * float(seconds1))
        time2 = str(tim[i + 1])
        hours2, minutes2, seconds2 = (["0", "0", "0"] + time2.split(":"))[-3:]

        miliseconds2 = int(3600000 * int(hours2) + 60000 * int(minutes2) + 1000 * float(seconds2))
        # print(miliseconds2-miliseconds1)
        hours3, milliseconds3 = divmod(miliseconds2 - miliseconds1, 3600000)
        minutes3, milliseconds3 = divmod(miliseconds2 - miliseconds1, 60000)
        seconds3 = float(miliseconds2 - miliseconds1) / 1000
        s2 = "%i:%02i:%06.3f" % (hours3, minutes3, seconds3)
        timelist.append(s2)
    # timelist.append(0)
    # print(timelist)
    # if i in timelist:
    timelist = timelist[0::2]
    # print(timelist) print("ds length")
    #print(len(ds))
    #print("timelist length")
    #print(len(timelist))
    leng = len(ds) / 2
    # print(ds.loc[2:3,"Message Type"])
    # if ds["Message Type"]=="Attach Accept":
    #     ds['Delay']=timelist


    for i in range(3, row_count + 1, 2):
        index = "C" + str(i)
        ws[index].fill = redFill
    loc_index=0
    for i in range(3, row_count + 1, 2):
        index = "H" + str(i)
        ws[index] = timelist[loc_index]
        loc_index = loc_index + 1

    wb.save('sample layer 3 format.xlsx')

    #ds_temp = pd.read_excel('sample layer 3 format.xlsx',sheet_name='L3 CSFBMO SR WITH DELAY CALC')

    avg = []
    for i in range(0, len(timelist) - 1):
        # print(str(tim[i+1]))

        time1 = str(timelist[i])
        hours1, minutes1, seconds1 = (["0", "0", "0"] + time1.split(":"))[-3:]
        miliseconds1 = int(3600000 * int(hours1) + 60000 * int(minutes1) + 1000 * float(seconds1))
        avg.append(miliseconds1)

    millis = 0

    for i in avg:
        millis = millis + i

    millis = millis / len(avg)

    seconds = (millis / 1000) % 60
    seconds = float(seconds)
    minutes = (millis / (1000 * 60)) % 60
    minutes = int(minutes)
    hours = (millis / (1000 * 60 * 60)) % 24
    print("%d:%d:%d" % (hours, minutes, seconds))
    avg_time = "%i:%02i:%06.3f" % (hours, minutes, seconds)
    print(avg_time)
    avg_time_scope = avg_time


    count_Extended_Service_Request = 0
    for i in ds['Message Type']:
        if i == 'Extended Service Request':
            count_Extended_Service_Request = count_Extended_Service_Request + 1


    for i in range(2, row_count + 1):
        ws["A" + str(i)].border = thin_border
        ws["B" + str(i)].border = thin_border
        ws["C" + str(i)].border = thin_border
        ws["D" + str(i)].border = thin_border
        ws["E" + str(i)].border = thin_border
        ws["F" + str(i)].border = thin_border
        ws["G" + str(i)].border = thin_border
        ws["H" + str(i)].border = thin_border

    wb.save('sample layer 3 format.xlsx')

Intrafreq_HOSR_with_delay_cal()